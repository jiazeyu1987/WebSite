from __future__ import annotations

import json
import mimetypes
import os
import re
import subprocess
import time
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse
from urllib.error import URLError
from urllib.request import Request, urlopen

from lxml import html
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill


BASE_URL = "https://int-medical.com"
TASK_DIR = Path(__file__).resolve().parent
ARTIFACTS_DIR = TASK_DIR / "artifacts"
IMAGES_DIR = ARTIFACTS_DIR / "images"
HALL_IMAGES_DIR = IMAGES_DIR / "halls"
PRODUCT_IMAGES_DIR = IMAGES_DIR / "products"
TMP_DIR = ARTIFACTS_DIR / "tmp"
OUTPUT_XLSX = ARTIFACTS_DIR / "int-medical-col14-showrooms.xlsx"

DEFAULT_NODE_EXECUTABLE = Path(
    os.environ.get(
        "CODEX_BUNDLED_NODE",
        r"C:\Users\BJB110\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe",
    )
)
DEFAULT_NODE_MODULES = Path(
    os.environ.get(
        "CODEX_BUNDLED_NODE_MODULES",
        r"C:\Users\BJB110\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\node_modules",
    )
)
DEFAULT_NODE_PNPM_MODULES = DEFAULT_NODE_MODULES / ".pnpm" / "node_modules"
SVG_RENDER_HELPER = TASK_DIR / "render_svg_preview.cjs"

COL22_SUBTYPES = (
    ("341", "接头类"),
    ("340", "阀类"),
    ("339", "金属丝加工及压力表"),
)

# User-approved fallback strings are intentionally explicit so downstream users
# can identify which exported cells came from source-data gaps.
DEFAULT_TEXT_FALLBACKS = {
    "product_description": "[FALLBACK] 源站缺少产品描述",
}


class ExtractionError(RuntimeError):
    """Fail-fast extraction error with a concrete precondition or field detail."""


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def absolute_url(value: str, base_url: str = BASE_URL) -> str:
    if not value:
        raise ExtractionError("Missing URL value required for extraction.")
    return urljoin(base_url, value)


def column_id_from_url(url: str) -> str:
    match = re.search(r"/col(\d+)/", url)
    if not match:
        raise ExtractionError(f"Cannot determine column id from hall URL: {url}")
    return match.group(1)


def product_id_from_url(url: str) -> str:
    match = re.search(r"/(\d+)$", url.rstrip("/"))
    if not match:
        raise ExtractionError(f"Cannot determine product id from detail URL: {url}")
    return match.group(1)


def request_bytes(url: str, *, payload: bytes | None = None, content_type: str | None = None) -> tuple[bytes, str]:
    headers = {"User-Agent": "Mozilla/5.0 Codex showroom extractor"}
    if content_type:
        headers["Content-Type"] = content_type
    request = Request(url, data=payload, headers=headers, method="POST" if payload is not None else "GET")
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            with urlopen(request, timeout=60) as response:
                body = response.read()
                response_content_type = response.headers.get("Content-Type", "")
            if not body:
                raise ExtractionError(f"Empty response body returned from {url}")
            return body, response_content_type
        except URLError as exc:
            last_error = exc
            if attempt == 3:
                break
            time.sleep(1)
    raise ExtractionError(f"Request failed for {url}: {last_error}") from last_error


def request_text(url: str) -> str:
    body, _ = request_bytes(url)
    return body.decode("utf-8", errors="strict")


def request_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    raw, _ = request_bytes(url, payload=body, content_type="application/json")
    try:
        return json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise ExtractionError(f"Invalid JSON returned from {url}: {exc}") from exc


def parse_showroom_navigation(page_html: str) -> list[dict[str, str]]:
    tree = html.fromstring(page_html)
    hall_nodes = tree.xpath('//div[contains(@class,"corpContentItem")]/a')
    rows: list[dict[str, str]] = []
    for hall_node in hall_nodes:
        hall_name = normalize_text("".join(hall_node.xpath('.//p[contains(@class,"corpContentItemText")]//text()')))
        hall_href = normalize_text("".join(hall_node.xpath("./@href")))
        icon_urls = [absolute_url(src) for src in hall_node.xpath('.//div[contains(@class,"corpContentItemimgBox")]//img/@src')]
        if not hall_name or not hall_href:
            raise ExtractionError("Missing hall name or hall URL while parsing showroom navigation.")
        if len(icon_urls) != 2:
            raise ExtractionError(
                f"Expected two icon URLs for hall `{hall_name}` but found {len(icon_urls)}."
            )
        hall_url = absolute_url(hall_href)
        rows.append(
            {
                "hall_name": hall_name,
                "hall_url": hall_url,
                "column_id": column_id_from_url(hall_url),
                "icon_default_url": icon_urls[0],
                "icon_active_url": icon_urls[1],
            }
        )
    if len(rows) != 7:
        raise ExtractionError(f"Expected 7 halls from showroom navigation but found {len(rows)}.")
    return rows


def _extract_total_pages(tree: html.HtmlElement, page_html: str) -> int:
    hrefs = tree.xpath('//a[contains(@href,"/list_")]/@href')
    page_numbers = [int(match.group(1)) for href in hrefs for match in [re.search(r"/list_(\d+)", href)] if match]
    if page_numbers:
        return max(page_numbers)
    if "/list_" in page_html:
        page_numbers = [int(match.group(1)) for match in re.finditer(r"/list_(\d+)", page_html)]
        if page_numbers:
            return max(page_numbers)
    return 1


def parse_html_product_list(
    page_html: str,
    *,
    hall_name: str,
    hall_url: str,
) -> tuple[list[dict[str, str]], int]:
    tree = html.fromstring(page_html)
    product_nodes = tree.xpath('//div[contains(@class,"corpListItem")]/a')
    rows: list[dict[str, str]] = []
    for product_node in product_nodes:
        detail_href = normalize_text("".join(product_node.xpath("./@href")))
        image_src = normalize_text("".join(product_node.xpath('.//div[contains(@class,"corpListItemImg")]//img/@src')))
        title = normalize_text("".join(product_node.xpath('.//p[contains(@class,"corpListItemIconsText")]//text()')))
        if not title:
            title = normalize_text("".join(product_node.xpath('.//img/@alt')))
        if not detail_href or not image_src or not title:
            raise ExtractionError(
                f"Missing detail URL, list image, or product title for hall `{hall_name}` in {hall_url}."
            )
        rows.append(
            {
                "hall_name": hall_name,
                "hall_url": hall_url,
                "subtype_name": "",
                "product_name": title,
                "product_detail_url": absolute_url(detail_href),
                "list_cover_url": absolute_url(image_src),
            }
        )
    if not rows:
        raise ExtractionError(f"No products found while parsing hall list page {hall_url}.")
    return rows, _extract_total_pages(tree, page_html)


def parse_col22_ajax_payload(
    payload: dict[str, Any],
    *,
    hall_name: str,
    hall_url: str,
    subtype_name: str,
) -> tuple[list[dict[str, str]], int]:
    if payload.get("success") is not True:
        raise ExtractionError(f"AJAX response for `{hall_name}` failed: {payload}")
    info = payload.get("info")
    if not isinstance(info, dict):
        raise ExtractionError(f"AJAX response for `{hall_name}` is missing `info`: {payload}")
    items = info.get("list")
    if not isinstance(items, list):
        raise ExtractionError(f"AJAX response for `{hall_name}` is missing `info.list`: {payload}")
    total_pages = int(info.get("totalPage") or 0)
    if total_pages < 1:
        raise ExtractionError(f"AJAX response for `{hall_name}` has invalid total pages: {payload}")

    rows: list[dict[str, str]] = []
    for item in items:
        title = normalize_text(str(item.get("TITLE", "")))
        detail_href = normalize_text(str(item.get("pageurl", "")))
        finfo = item.get("finfo") or {}
        image_src = normalize_text(str(finfo.get("COVERIMAGE", "")))
        if not title or not detail_href or not image_src:
            raise ExtractionError(
                f"AJAX product item for `{hall_name}` / `{subtype_name}` is missing title, detail URL, or cover image: {item}"
            )
        rows.append(
            {
                "hall_name": hall_name,
                "hall_url": hall_url,
                "subtype_name": subtype_name,
                "product_name": title,
                "product_detail_url": absolute_url(detail_href),
                "list_cover_url": absolute_url(image_src),
            }
        )
    if not rows:
        raise ExtractionError(f"AJAX hall `{hall_name}` / `{subtype_name}` returned no products.")
    return rows, total_pages


def parse_product_detail(page_html: str) -> dict[str, str]:
    tree = html.fromstring(page_html)
    product_name = normalize_text("".join(tree.xpath('//div[contains(@class,"productTopcontent-view")]//h2//text()')))
    description = normalize_text("".join(tree.xpath('(//div[contains(@class,"productTopcontent-view")]//p)[1]//text()')))
    image_src = normalize_text("".join(tree.xpath('(//div[contains(@class,"productTopImg")]//img/@src)[1]')))
    missing_fields: list[str] = []
    if not product_name:
        missing_fields.append("product title")
    if not image_src:
        missing_fields.append("primary image")
    if missing_fields:
        raise ExtractionError(
            "Product detail page is missing required field(s): " + ", ".join(missing_fields) + "."
        )
    if not description:
        description = DEFAULT_TEXT_FALLBACKS["product_description"]

    feature_candidates = tree.xpath(
        '//div[contains(@class,"productCenterContentLeft")]//li | //div[contains(@class,"productCenterContentLeft")]//p'
    )
    feature_lines: list[str] = []
    for node in feature_candidates:
        line = normalize_text("".join(node.xpath(".//text()")))
        if line and line != description and line not in feature_lines:
            feature_lines.append(line)

    return {
        "product_name": product_name,
        "product_description": description,
        "product_image_url": absolute_url(image_src),
        "product_features": "\n".join(feature_lines),
    }


def ensure_directory(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def guess_extension(url: str, content_type: str) -> str:
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix:
        return suffix
    guessed = mimetypes.guess_extension((content_type or "").split(";")[0].strip())
    if guessed:
        return guessed
    return ".bin"


def download_file(url: str, destination: Path) -> Path:
    raw, content_type = request_bytes(url)
    extension = guess_extension(url, content_type)
    output_path = destination.with_suffix(extension)
    ensure_directory(output_path.parent)
    output_path.write_bytes(raw)
    return output_path


def convert_svg_to_png(svg_path: Path, png_path: Path, *, size_px: int = 96) -> Path:
    if not DEFAULT_NODE_EXECUTABLE.exists():
        raise ExtractionError(
            f"SVG preview conversion prerequisite is missing: bundled node executable not found at {DEFAULT_NODE_EXECUTABLE}"
        )
    if not DEFAULT_NODE_MODULES.exists():
        raise ExtractionError(
            f"SVG preview conversion prerequisite is missing: bundled node_modules not found at {DEFAULT_NODE_MODULES}"
        )
    ensure_directory(png_path.parent)
    env = dict(os.environ)
    node_paths = [str(DEFAULT_NODE_MODULES)]
    if DEFAULT_NODE_PNPM_MODULES.exists():
        node_paths.append(str(DEFAULT_NODE_PNPM_MODULES))
    env["NODE_PATH"] = os.pathsep.join(node_paths)
    result = subprocess.run(
        [str(DEFAULT_NODE_EXECUTABLE), str(SVG_RENDER_HELPER), str(svg_path), str(png_path), str(size_px)],
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )
    if result.returncode != 0:
        raise ExtractionError(
            "SVG preview conversion failed for "
            f"{svg_path}: {result.stderr.strip() or result.stdout.strip() or 'unknown sharp failure'}"
        )
    if not png_path.exists():
        raise ExtractionError(f"SVG preview conversion did not create expected PNG file: {png_path}")
    return png_path


def prepare_hall_preview(icon_default_path: Path, preview_dir: Path) -> Path:
    preview_path = preview_dir / f"{icon_default_path.stem}-preview.png"
    if icon_default_path.suffix.lower() == ".svg":
        return convert_svg_to_png(icon_default_path, preview_path)
    return icon_default_path


def build_html_page_url(hall_url: str, page_number: int) -> str:
    if page_number == 1:
        return hall_url
    return re.sub(r"/list$", f"/list_{page_number}", hall_url)


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_preview_image(sheet, image_path: str, cell_ref: str, *, width: int, height: int) -> None:
    if not image_path:
        raise ExtractionError(f"Missing local preview image path required for workbook cell {cell_ref}.")
    file_path = Path(image_path)
    if not file_path.exists():
        raise ExtractionError(f"Workbook preview image file does not exist: {file_path}")
    image = ExcelImage(str(file_path))
    image.width = width
    image.height = height
    sheet.add_image(image, cell_ref)


def build_workbook(output_path: Path, hall_rows: list[dict[str, str]], product_rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    hall_sheet = workbook.active
    hall_sheet.title = "展厅"
    product_sheet = workbook.create_sheet("产品")

    _write_header(
        hall_sheet,
        ["展厅名称", "栏目URL", "ICON预览", "ICON默认本地路径", "ICON默认源URL", "ICON高亮本地路径", "ICON高亮源URL"],
    )
    _write_header(
        product_sheet,
        ["展厅名称", "展厅栏目URL", "二级分类", "产品名称", "产品详情URL", "产品图预览", "产品图片本地路径", "产品图片源URL", "产品描述", "产品特点"],
    )

    hall_sheet.freeze_panes = "A2"
    product_sheet.freeze_panes = "A2"
    hall_sheet.column_dimensions["A"].width = 18
    hall_sheet.column_dimensions["B"].width = 34
    hall_sheet.column_dimensions["C"].width = 16
    hall_sheet.column_dimensions["D"].width = 42
    hall_sheet.column_dimensions["E"].width = 44
    hall_sheet.column_dimensions["F"].width = 42
    hall_sheet.column_dimensions["G"].width = 44
    for column in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
        product_sheet.column_dimensions[column].width = 20
    product_sheet.column_dimensions["B"].width = 34
    product_sheet.column_dimensions["E"].width = 36
    product_sheet.column_dimensions["F"].width = 16
    product_sheet.column_dimensions["G"].width = 42
    product_sheet.column_dimensions["H"].width = 44
    product_sheet.column_dimensions["I"].width = 36
    product_sheet.column_dimensions["J"].width = 42

    for index, row in enumerate(hall_rows, start=2):
        hall_sheet.append(
            [
                row["hall_name"],
                row["hall_url"],
                "",
                row["icon_default_path"],
                row["icon_default_url"],
                row["icon_active_path"],
                row["icon_active_url"],
            ]
        )
        hall_sheet.row_dimensions[index].height = 46
        preview_path = row.get("icon_preview_path") or row["icon_default_path"]
        _add_preview_image(hall_sheet, preview_path, f"C{index}", width=34, height=34)

    for index, row in enumerate(product_rows, start=2):
        product_sheet.append(
            [
                row["hall_name"],
                row["hall_url"],
                row["subtype_name"],
                row["product_name"],
                row["product_detail_url"],
                "",
                row["product_image_path"],
                row["product_image_url"],
                row["product_description"],
                row["product_features"],
            ]
        )
        product_sheet.row_dimensions[index].height = 60
        _add_preview_image(product_sheet, row["product_image_path"], f"F{index}", width=48, height=48)

    for sheet in (hall_sheet, product_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ensure_directory(output_path.parent)
    workbook.save(output_path)


def extract_hall_products(hall: dict[str, str]) -> list[dict[str, str]]:
    hall_name = hall["hall_name"]
    hall_url = hall["hall_url"]
    column_id = hall["column_id"]
    if column_id == "22":
        return extract_col22_products(hall_name, hall_url)

    first_page_html = request_text(hall_url)
    first_page_rows, total_pages = parse_html_product_list(first_page_html, hall_name=hall_name, hall_url=hall_url)
    all_rows = list(first_page_rows)
    for page_number in range(2, total_pages + 1):
        page_url = build_html_page_url(hall_url, page_number)
        page_html = request_text(page_url)
        page_rows, _ = parse_html_product_list(page_html, hall_name=hall_name, hall_url=hall_url)
        all_rows.extend(page_rows)
    return all_rows


def extract_col22_products(hall_name: str, hall_url: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    ajax_base = f"{BASE_URL}/ajax/list_page?siteId=1&columnId=22&subId=0&pageNumber={{page}}&pageSize=9"
    for subtype_id, subtype_name in COL22_SUBTYPES:
        page_number = 0
        total_pages: int | None = None
        while total_pages is None or page_number < total_pages:
            payload = request_json(
                ajax_base.format(page=page_number),
                {"EQ_FENLEI": f",{subtype_id},"},
            )
            page_rows, total_pages = parse_col22_ajax_payload(
                payload,
                hall_name=hall_name,
                hall_url=hall_url,
                subtype_name=subtype_name,
            )
            rows.extend(page_rows)
            page_number += 1
    return rows


def enrich_product(product_row: dict[str, str]) -> dict[str, str]:
    detail_html = request_text(product_row["product_detail_url"])
    try:
        detail = parse_product_detail(detail_html)
    except ExtractionError as exc:
        raise ExtractionError(
            "Product detail extraction failed for "
            f"hall `{product_row['hall_name']}`, product `{product_row['product_name']}`, "
            f"url `{product_row['product_detail_url']}`: {exc}"
        ) from exc
    product_id = product_id_from_url(product_row["product_detail_url"])
    column_id = column_id_from_url(product_row["product_detail_url"])
    product_dir = PRODUCT_IMAGES_DIR / column_id
    image_path = download_file(detail["product_image_url"], product_dir / product_id)
    merged = dict(product_row)
    merged.update(detail)
    merged["product_image_path"] = str(image_path)
    return merged


def enrich_hall_assets(hall_row: dict[str, str]) -> dict[str, str]:
    hall_dir = HALL_IMAGES_DIR / hall_row["column_id"]
    default_path = download_file(hall_row["icon_default_url"], hall_dir / "icon-default")
    active_path = download_file(hall_row["icon_active_url"], hall_dir / "icon-active")
    preview_path = prepare_hall_preview(default_path, TMP_DIR / "hall-previews" / hall_row["column_id"])
    enriched = dict(hall_row)
    enriched["icon_default_path"] = str(default_path)
    enriched["icon_active_path"] = str(active_path)
    enriched["icon_preview_path"] = str(preview_path)
    return enriched


def run() -> dict[str, int]:
    ensure_directory(ARTIFACTS_DIR)
    ensure_directory(HALL_IMAGES_DIR)
    ensure_directory(PRODUCT_IMAGES_DIR)
    ensure_directory(TMP_DIR)

    root_html = request_text(f"{BASE_URL}/col14/list")
    hall_rows = [enrich_hall_assets(row) for row in parse_showroom_navigation(root_html)]
    product_rows: list[dict[str, str]] = []
    for hall_row in hall_rows:
        hall_products = extract_hall_products(hall_row)
        for product_row in hall_products:
            product_rows.append(enrich_product(product_row))

    build_workbook(OUTPUT_XLSX, hall_rows, product_rows)
    return {"hall_count": len(hall_rows), "product_count": len(product_rows)}


def main() -> None:
    summary = run()
    print(
        json.dumps(
            {
                "workbook": str(OUTPUT_XLSX),
                "hall_count": summary["hall_count"],
                "product_count": summary["product_count"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
