from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from urllib.error import URLError
from unittest.mock import patch

from openpyxl import load_workbook
from PIL import Image as PilImage

from extract_showrooms import (
    build_workbook,
    parse_col22_ajax_payload,
    parse_html_product_list,
    parse_product_detail,
    parse_showroom_navigation,
    request_bytes,
)


TASK_DIR = Path(__file__).resolve().parent
FIXTURES_DIR = TASK_DIR / "fixtures"


def read_fixture(name: str) -> str:
    return (FIXTURES_DIR / name).read_text(encoding="utf-8")


class ParseShowroomNavigationTests(unittest.TestCase):
    def test_parses_all_seven_showrooms_with_dual_icons(self) -> None:
        rows = parse_showroom_navigation(read_fixture("col14_showrooms.html"))

        self.assertEqual(7, len(rows))
        self.assertEqual("心内介入类", rows[0]["hall_name"])
        self.assertEqual("https://int-medical.com/col14/list", rows[0]["hall_url"])
        self.assertEqual(
            "https://int-medical.com/upload/file/2024-01/col15/1706065284308.svg",
            rows[0]["icon_default_url"],
        )
        self.assertEqual(
            "https://int-medical.com/upload/file/2024-03/col15/1710065623989.svg",
            rows[-1]["icon_active_url"],
        )


class ParseHtmlProductListTests(unittest.TestCase):
    def test_parses_products_and_total_pages_from_html_hall(self) -> None:
        rows, total_pages = parse_html_product_list(
            read_fixture("col14_list_page.html"),
            hall_name="心内介入类",
            hall_url="https://int-medical.com/col14/list",
        )

        self.assertEqual(3, total_pages)
        self.assertEqual(2, len(rows))
        self.assertEqual("心内介入类", rows[0]["hall_name"])
        self.assertEqual("导管鞘组", rows[0]["product_name"])
        self.assertEqual("https://int-medical.com/col14/3503", rows[0]["product_detail_url"])
        self.assertEqual(
            "https://int-medical.com/upload/image/2024-02/col14/1708248438056.png",
            rows[1]["list_cover_url"],
        )


class ParseAjaxHallTests(unittest.TestCase):
    def test_parses_ajax_products_with_subtype_and_paging(self) -> None:
        payload = json.loads(read_fixture("col22_ajax_page.json"))

        rows, total_pages = parse_col22_ajax_payload(
            payload,
            hall_name="医疗标准件",
            hall_url="https://int-medical.com/col22/list",
            subtype_name="接头类",
        )

        self.assertEqual(2, total_pages)
        self.assertEqual(2, len(rows))
        self.assertEqual("接头类", rows[0]["subtype_name"])
        self.assertEqual("输液接头及附件", rows[0]["product_name"])
        self.assertEqual("https://int-medical.com/col22/3031", rows[0]["product_detail_url"])
        self.assertEqual(
            "https://int-medical.com/upload/image/2024-02/col22/1708307865415.png",
            rows[1]["list_cover_url"],
        )


class ParseProductDetailTests(unittest.TestCase):
    def test_parses_description_primary_image_and_features(self) -> None:
        row = parse_product_detail(read_fixture("detail_with_features.html"))

        self.assertEqual("导管鞘组", row["product_name"])
        self.assertEqual(
            "适用于建立有助于血管内器械的经皮进入通道。",
            row["product_description"],
        )
        self.assertEqual(
            "https://int-medical.com/upload/image/2025-10/col14/1761801848324.png",
            row["product_image_url"],
        )
        self.assertEqual(
            "本产品采用旋钮式止血阀。\n超薄管壁，同尺寸外径下腔道更大。\n规格范围广。",
            row["product_features"],
        )

    def test_allows_missing_features_block(self) -> None:
        row = parse_product_detail(read_fixture("detail_without_features.html"))

        self.assertEqual("输液接头及附件", row["product_name"])
        self.assertEqual("用于注射液体、重力输液、液体抽取。", row["product_description"])
        self.assertEqual("", row["product_features"])

    def test_uses_default_string_when_description_is_missing(self) -> None:
        row = parse_product_detail(read_fixture("detail_missing_description.html"))

        self.assertEqual("旋塞阀", row["product_name"])
        self.assertEqual("[FALLBACK] 源站缺少产品描述", row["product_description"])
        self.assertEqual(
            "45/500/1000psi以下无泄漏\n转动平稳，使用方便\n麻醉药专用耐脂材料",
            row["product_features"],
        )


class WorkbookSmokeTests(unittest.TestCase):
    def test_builds_workbook_with_expected_sheets_and_embedded_images(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image_path = tmp_path / "preview.png"
            PilImage.new("RGB", (32, 32), color=(16, 128, 255)).save(image_path)
            output_path = tmp_path / "showrooms.xlsx"

            hall_rows = [
                {
                    "hall_name": "心内介入类",
                    "hall_url": "https://int-medical.com/col14/list",
                    "icon_default_path": str(image_path),
                    "icon_default_url": "https://int-medical.com/upload/file/example-default.svg",
                    "icon_active_path": str(image_path),
                    "icon_active_url": "https://int-medical.com/upload/file/example-active.svg",
                }
            ]
            product_rows = [
                {
                    "hall_name": "心内介入类",
                    "hall_url": "https://int-medical.com/col14/list",
                    "subtype_name": "",
                    "product_name": "导管鞘组",
                    "product_detail_url": "https://int-medical.com/col14/3503",
                    "product_image_path": str(image_path),
                    "product_image_url": "https://int-medical.com/upload/image/example.png",
                    "product_description": "适用于建立有助于血管内器械的经皮进入通道。",
                    "product_features": "规格范围广。",
                }
            ]

            build_workbook(output_path, hall_rows, product_rows)

            workbook = load_workbook(output_path)
            self.assertEqual(["展厅", "产品"], workbook.sheetnames)
            self.assertEqual("展厅名称", workbook["展厅"]["A1"].value)
            self.assertEqual("产品名称", workbook["产品"]["D1"].value)
            self.assertGreater(len(workbook["展厅"]._images), 0)
            self.assertGreater(len(workbook["产品"]._images), 0)


class RequestBytesRetryTests(unittest.TestCase):
    @patch("extract_showrooms.time.sleep", return_value=None)
    @patch("extract_showrooms.urlopen")
    def test_retries_transient_transport_errors(self, mocked_urlopen, _sleep) -> None:
        successful_response = mocked_urlopen.return_value
        successful_response.__enter__.return_value.read.return_value = b"ok"
        successful_response.__enter__.return_value.headers.get.return_value = "text/plain"
        mocked_urlopen.side_effect = [URLError("temporary eof"), successful_response]

        body, content_type = request_bytes("https://example.com/demo")

        self.assertEqual(b"ok", body)
        self.assertEqual("text/plain", content_type)
        self.assertEqual(2, mocked_urlopen.call_count)


if __name__ == "__main__":
    unittest.main()
